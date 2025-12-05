"""Service for generating Excel exports of regulation requirements."""
from io import BytesIO
from typing import List, Dict, Any, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime


def create_regulation_worksheet(workbook: Workbook, regulation: Dict[str, Any]) -> None:
    """Create a worksheet for a single regulation."""
    # Create worksheet with sanitized name (Excel sheet names have restrictions)
    sheet_name = regulation.get('friendly_name', 'Unnamed Regulation')[:31]  # Max 31 chars
    # Remove invalid characters
    invalid_chars = ['\\', '/', '?', '*', '[', ']', ':']
    for char in invalid_chars:
        sheet_name = sheet_name.replace(char, '_')
    
    ws = workbook.create_sheet(title=sheet_name)
    
    # Styles
    header_font = Font(bold=True, size=12)
    section_font = Font(bold=True, size=11)
    field_name_font = Font(bold=False, size=10)
    description_font = Font(size=9, italic=True)
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    section_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    alt_row_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    header_font_white = Font(bold=True, size=12, color="FFFFFF")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    row = 1
    
    # Header Section
    ws.merge_cells(f'A{row}:C{row}')
    header_cell = ws[f'A{row}']
    header_cell.value = regulation.get('friendly_name', 'Unnamed Regulation')
    header_cell.font = header_font_white
    header_cell.fill = header_fill
    header_cell.alignment = Alignment(horizontal='left', vertical='center')
    row += 1
    
    # Country, Number Type, End User Type
    ws.merge_cells(f'A{row}:C{row}')
    info_cell = ws[f'A{row}']
    info_parts = []
    if regulation.get('iso_country'):
        info_parts.append(f"Country: {regulation['iso_country']}")
    if regulation.get('number_type'):
        info_parts.append(f"Number Type: {regulation['number_type']}")
    if regulation.get('end_user_type'):
        info_parts.append(f"End User Type: {regulation['end_user_type']}")
    info_cell.value = " | ".join(info_parts) if info_parts else "Regulation Information"
    info_cell.font = Font(size=10)
    info_cell.alignment = Alignment(horizontal='left', vertical='center')
    row += 1
    
    # Last Updated
    ws.merge_cells(f'A{row}:C{row}')
    updated_cell = ws[f'A{row}']
    last_updated = regulation.get('last_updated')
    if last_updated:
        if isinstance(last_updated, str):
            updated_cell.value = f"Last Updated: {last_updated}"
        else:
            updated_cell.value = f"Last Updated: {last_updated.strftime('%Y-%m-%d %H:%M:%S')}"
    else:
        updated_cell.value = "Last Updated: N/A"
    updated_cell.font = Font(size=9, italic=True)
    updated_cell.alignment = Alignment(horizontal='left', vertical='center')
    row += 2
    
    requirements = regulation.get('requirements', {})
    
    # End User Requirements Section
    end_user_reqs = requirements.get('end_user', [])
    if end_user_reqs:
        # Section Header
        ws.merge_cells(f'A{row}:C{row}')
        section_cell = ws[f'A{row}']
        section_cell.value = "Required Information"
        section_cell.font = section_font
        section_cell.fill = section_fill
        section_cell.alignment = Alignment(horizontal='left', vertical='center')
        section_cell.border = thin_border
        row += 1
        
        # Table Header
        ws[f'A{row}'] = "Field Name"
        ws[f'B{row}'] = "Description"
        ws[f'C{row}'] = "Value"
        for col in ['A', 'B', 'C']:
            cell = ws[f'{col}{row}']
            cell.font = header_font
            cell.fill = header_fill
            cell.font = header_font_white
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = thin_border
        row += 1
        
        # Process each end user requirement
        for end_user_req in end_user_reqs:
            detailed_fields = end_user_req.get('detailed_fields', [])
            for idx, field in enumerate(detailed_fields):
                ws[f'A{row}'] = field.get('friendly_name', field.get('machine_name', ''))
                ws[f'B{row}'] = field.get('description', '')
                ws[f'C{row}'] = ""  # Empty value column for user input
                
                # Apply styling
                for col in ['A', 'B', 'C']:
                    cell = ws[f'{col}{row}']
                    cell.border = thin_border
                    if idx % 2 == 1:  # Alternate row colors
                        cell.fill = alt_row_fill
                    if col == 'A':
                        cell.font = field_name_font
                    elif col == 'B':
                        cell.font = description_font
                    else:
                        cell.font = Font(size=10)
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                row += 1
        
        row += 1
    
    # Supporting Document Requirements Section
    supporting_docs = requirements.get('supporting_document', [])
    if supporting_docs:
        # Flatten nested arrays
        flattened_docs = []
        for doc_group in supporting_docs:
            if isinstance(doc_group, list):
                flattened_docs.extend(doc_group)
            else:
                flattened_docs.append(doc_group)
        
        if flattened_docs:
            # Section Header
            ws.merge_cells(f'A{row}:C{row}')
            section_cell = ws[f'A{row}']
            section_cell.value = "Required Documents"
            section_cell.font = section_font
            section_cell.fill = section_fill
            section_cell.alignment = Alignment(horizontal='left', vertical='center')
            section_cell.border = thin_border
            row += 1
            
            # Process each document requirement
            for doc_req in flattened_docs:
                # Document Name
                ws.merge_cells(f'A{row}:C{row}')
                doc_name_cell = ws[f'A{row}']
                doc_name_cell.value = doc_req.get('name', 'Document Requirement')
                doc_name_cell.font = Font(bold=True, size=11)
                doc_name_cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
                doc_name_cell.alignment = Alignment(horizontal='left', vertical='center')
                doc_name_cell.border = thin_border
                row += 1
                
                # Description
                if doc_req.get('description'):
                    ws.merge_cells(f'A{row}:C{row}')
                    desc_cell = ws[f'A{row}']
                    desc_cell.value = doc_req.get('description')
                    desc_cell.font = Font(size=9, italic=True)
                    desc_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    row += 1
                
                # Accepted Documents
                accepted_docs = doc_req.get('accepted_documents', [])
                if accepted_docs:
                    # Add instruction if multiple document types are accepted
                    if len(accepted_docs) > 1:
                        ws.merge_cells(f'A{row}:C{row}')
                        instruction_cell = ws[f'A{row}']
                        instruction_cell.value = "Choose ONE of the following document types:"
                        instruction_cell.font = Font(bold=True, size=10, italic=True)
                        instruction_cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                        instruction_cell.alignment = Alignment(horizontal='left', vertical='center')
                        instruction_cell.border = thin_border
                        row += 1
                    
                    for accepted_doc in accepted_docs:
                        # Document Type Name
                        ws.merge_cells(f'A{row}:C{row}')
                        doc_type_cell = ws[f'A{row}']
                        doc_type_cell.value = f"  Accepted: {accepted_doc.get('name', 'Document')}"
                        doc_type_cell.font = Font(bold=True, size=10)
                        doc_type_cell.alignment = Alignment(horizontal='left', vertical='center')
                        row += 1
                        
                        # Fields for this document type
                        detailed_fields = accepted_doc.get('detailed_fields', [])
                        if detailed_fields:
                            # Sub-header for fields - only Field Name column
                            ws[f'A{row}'] = "Field Name"
                            cell = ws[f'A{row}']
                            cell.font = Font(bold=True, size=9)
                            cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                            cell.alignment = Alignment(horizontal='left', vertical='center')
                            cell.border = thin_border
                            row += 1
                            
                            # Field rows - only Field Name column
                            for idx, field in enumerate(detailed_fields):
                                ws[f'A{row}'] = field.get('friendly_name', field.get('machine_name', ''))
                                
                                cell = ws[f'A{row}']
                                cell.border = thin_border
                                if idx % 2 == 1:
                                    cell.fill = alt_row_fill
                                cell.font = field_name_font
                                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                                row += 1
                        else:
                            # No specific fields, just document type
                            ws.merge_cells(f'A{row}:C{row}')
                            no_fields_cell = ws[f'A{row}']
                            no_fields_cell.value = "  (No specific fields required)"
                            no_fields_cell.font = Font(size=9, italic=True)
                            no_fields_cell.alignment = Alignment(horizontal='left', vertical='center')
                            row += 1
                
                row += 1  # Space between document requirements
    
    # Auto-adjust column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 30
    
    # Freeze header row (first 3 rows are header info)
    ws.freeze_panes = 'A4'


def generate_regulation_excel(regulations: List[Dict[str, Any]]) -> BytesIO:
    """Generate an Excel workbook from one or more regulations.
    
    Args:
        regulations: List of regulation dictionaries with all required fields
        
    Returns:
        BytesIO object containing the Excel file
    """
    workbook = Workbook()
    # Remove default sheet
    if 'Sheet' in workbook.sheetnames:
        workbook.remove(workbook['Sheet'])
    
    # Create a worksheet for each regulation
    for regulation in regulations:
        create_regulation_worksheet(workbook, regulation)
    
    # Save to BytesIO
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    
    return output

